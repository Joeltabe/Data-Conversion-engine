#!/usr/bin/env python3
"""
FORM B — Department course catalogue cross-check & generation.

Reads an authoritative per-department/level course list (e.g. the
"FORM B LEVEL 300 NURSING.xlsx" workbook), cross-checks it against the
courses extracted from student transcripts, and emits a FORM B workbook
with the exact column layout:

    Course Code | Descriptions | credit | Semester | Department |
    Levels | School id | year | section
"""

import re
import os
from collections import Counter

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    raise SystemExit("ERROR: openpyxl not installed. Run: pip install openpyxl")

DEFAULT_SECTION = 1
DEFAULT_SCHOOL_ID = 10

# Matricule prefix (letters between the LMU/LMUI prefix and the digits)
# -> canonical department name. Keep in sync with the Department column
# used inside the FORM B catalogues.
DEPARTMENT_MAP = {
    "NUR": "NURSING",
    "MID": "MIDWIFERY",
    "ACC": "ACCOUNTING",
    "AGP": "AGRICULTURAL AND PRODUCTION TECHNOLOGY",
    "SWE": "SOFTWARE ENGINEERING",
    "NET": "NETWORK AND SECURITY",
    "CVE": "CIVIL ENGINEERING",
    "MEC": "MECHANICAL ENGINEERING",
    "ELT": "ELECTRICAL ENGINEERING",
    "LAW": "LAW",
    "HOS": "HOSPITALITY MANAGEMENT",
    "BIO": "BIOSCIENCES",
    "BSC": "BIOSCIENCES",
}

# ── colours (mirror the engine palette) ──────────────────────────────────────
DARK_BLUE = "1F4E79"
PALE_BLUE = "EBF3FB"
PALE_YELL = "FFF2CC"
PALE_GREY = "F8F9FA"
PALE_GREEN = "E2EFDA"
RED_FILL = "FFE0E0"
TITLE_BG = "D6E4F0"
thin = Side(style="thin", color="BFBFBF")
bdr = Border(left=thin, right=thin, top=thin, bottom=thin)

SEM_FILLS = {1: PALE_BLUE, 2: PALE_YELL, 3: "E8F5E9", 4: "F3E5F5"}

FORM_B_HEADERS = [
    "Course Code", "Descriptions", "credit", "Semester",
    "Department", "Levels", "School id", "year", "section",
]


def _f(val, default=0.0):
    try:
        return float(str(val).replace(",", ".").strip())
    except Exception:
        return default


def normalize_code(code):
    """Upper-case, space-collapsed course code for display comparisons."""
    return re.sub(r"\s+", "", str(code or "")).upper()


def code_key(code):
    """OCR-tolerant course code key (O->0, I->1, L->1) for matching."""
    k = normalize_code(code)
    for a, b in (("O", "0"), ("I", "1"), ("L", "1")):
        k = k.replace(a, b)
    return k


# ═══════════════════════════════════════════════════════════════════════════
#  CATALOGUE LOADING
# ═══════════════════════════════════════════════════════════════════════════
def load_form_b(path):
    """Read a FORM B xlsx and return a list of course row dicts."""
    wb = openpyxl.load_workbook(str(path), data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    header = [str(h).strip().lower() if h is not None else "" for h in rows[0]]
    idx = {name: i for i, name in enumerate(header)}

    def col(name, row, default=""):
        i = idx.get(name)
        return row[i] if i is not None and i < len(row) and row[i] is not None else default

    result = []
    for r in rows[1:]:
        code = col("course code", r) or col("code", r)
        if not code:
            continue
        result.append({
            "code": str(code).strip(),
            "description": str(col("descriptions", r) or col("description", r) or "").strip(),
            "credit": _f(col("credit", r)),
            "semester": int(_f(col("semester", r), 1)),
            "department": str(col("department", r, "")).strip().upper(),
            "levels": int(_f(col("levels", r) or col("level", r))),
            "school_id": int(_f(col("school id", r, DEFAULT_SCHOOL_ID))),
            "year": str(col("year", r, "")).strip(),
            "section": int(_f(col("section", r, DEFAULT_SECTION), DEFAULT_SECTION)),
        })
    return result


def detect_department(matricule):
    """Infer the department from a matricule's letter prefix (NUR->NURSING)."""
    m = re.sub(r"\s+", "", str(matricule or ""))
    m = re.sub(r"^LMU[\-I]*", "", m, flags=re.IGNORECASE)
    m = re.sub(r"^\d+", "", m)
    match = re.match(r"^([A-Z]+)", m, flags=re.IGNORECASE)
    if not match:
        return ""
    prefix = match.group(1).upper()
    return DEPARTMENT_MAP.get(prefix, "")


def majority_department(matricules):
    """Most common department across a set of matricules (empty if none)."""
    votes = Counter(detect_department(m) for m in matricules)
    votes.pop("", None)
    return votes.most_common(1)[0][0] if votes else ""


def _spec_of(s):
    v = s.get("specialty") if isinstance(s, dict) else getattr(s, "specialty", "")
    return (v or "").strip()


def _mat_of(s):
    v = s.get("matricule") if isinstance(s, dict) else getattr(s, "matricule", "")
    return (v or "").strip()


def majority_specialty(students):
    """Most common specialty across students (PDF header 'SPECIALTY:')."""
    votes = Counter(_spec_of(s).upper() for s in students)
    votes.pop("", None)
    return votes.most_common(1)[0][0] if votes else ""


def infer_department(students):
    """Department from the transcript's SPECIALTY line (authoritative),
    falling back to the matricule letter prefix when absent."""
    sp = majority_specialty(students)
    if sp:
        return sp
    return majority_department([_mat_of(s) for s in students])


def _norm_dept(text):
    return re.sub(r"[^A-Z0-9]", "", str(text or "").upper())


def parse_catalog_filename(filename):
    """Extract (department, level) from a catalogue file name.

    e.g. "FORM B LEVEL 300 NURSING.xlsx" -> ("NURSING", 300)
    """
    base = os.path.splitext(os.path.basename(filename))[0]
    lm = re.search(r"LEVEL[\s_\-]*(\d{2,3})", base, flags=re.IGNORECASE)
    level = int(lm.group(1)) if lm else 0
    dept = re.sub(r"(?i)FORM\s*B|LEVEL[\s_\-]*\d{2,3}", " ", base)
    dept = re.sub(r"[\s_\-]+", " ", dept).strip().upper()
    return dept, level


def find_catalogs(form_b_dir, department="", level=0):
    """List every catalogue in the folder with its parsed dept/level.

    When department/level are given, catalogues matching them are returned
    first, in their folder order.
    """
    entries = []
    if not os.path.isdir(form_b_dir):
        return entries
    for name in sorted(os.listdir(form_b_dir)):
        if not name.lower().endswith(".xlsx"):
            continue
        dept, lvl = parse_catalog_filename(name)
        entries.append({
            "name": name,
            "path": os.path.join(form_b_dir, name),
            "department": dept,
            "level": lvl,
        })

    want_dept = _norm_dept(department)
    def rank(e):
        hits = 0
        if want_dept and e["department"] and (
                _norm_dept(e["department"]) == want_dept
                or _norm_dept(e["department"]) in want_dept
                or want_dept in _norm_dept(e["department"])):
            hits += 2
        if level and e["level"] == level:
            hits += 1
        return -hits

    return sorted(entries, key=rank)


# ═══════════════════════════════════════════════════════════════════════════
#  CROSS-CHECK
# ═══════════════════════════════════════════════════════════════════════════
def cross_check(students, catalog):
    """Compare every student's courses against the catalogue.

    students: list of engine.StudentRecord objects.
    Returns a dict with per-catalogue-course coverage and per-student
    missing / unexpected / mismatched course reports.
    """
    cat_by_key = {}
    for r in catalog:
        cat_by_key.setdefault(code_key(r["code"]), []).append(r)

    course_stats = []
    for r in catalog:
        key = code_key(r["code"])
        takers = []
        for s in students:
            if any(code_key(c.code) == key for c in s.courses):
                takers.append(s.matricule)
        course_stats.append({
            **r,
            "taken_by": takers,
            "count": len(takers),
        })

    per_student = []
    unexpected_agg = {}
    for s in students:
        s_by_key = {code_key(c.code): c for c in s.courses}
        missing = []
        for r in catalog:
            if code_key(r["code"]) not in s_by_key:
                missing.append(r)
        unexpected = []
        issues = []
        for key, c in s_by_key.items():
            if key in cat_by_key:
                cr = cat_by_key[key][0]
                if cr["credit"] and cr["credit"] != c.credit_value:
                    issues.append({
                        "code": c.code, "type": "credit_mismatch",
                        "student_credit": c.credit_value,
                        "catalog_credit": cr["credit"],
                    })
                if cr["semester"] != c.semester:
                    issues.append({
                        "code": c.code, "type": "semester_mismatch",
                        "student_semester": c.semester,
                        "catalog_semester": cr["semester"],
                    })
            else:
                unexpected.append({
                    "code": c.code, "title": c.title,
                    "semester": c.semester, "credit": c.credit_value,
                })
                unexpected_agg.setdefault(code_key(c.code), {
                    "code": c.code, "title": c.title,
                    "semester": c.semester, "credit": c.credit_value,
                    "students": [],
                })["students"].append(s.matricule)

        per_student.append({
            "matricule": s.matricule,
            "name": s.name,
            "taken": len(s_by_key),
            "matched": len(s_by_key) - len(unexpected),
            "missing": missing,
            "unexpected": unexpected,
            "issues": issues,
        })

    return {
        "catalog_count": len(catalog),
        "student_count": len(students),
        "course_stats": course_stats,
        "students": per_student,
        "unexpected": list(unexpected_agg.values()),
    }


# ═══════════════════════════════════════════════════════════════════════════
#  FORM B DERIVED FROM UPLOADED TRANSCRIPTS
# ═══════════════════════════════════════════════════════════════════════════
def derive_form_b_from_students(students, department="", level="", year="",
                                school_id=DEFAULT_SCHOOL_ID, section=DEFAULT_SECTION):
    """Build the FORM B course list from the uploaded transcripts.

    The same course taken by several students is deduplicated (OCR-tolerant)
    and the most common code / description / credit / semester across the
    cohort is kept as the canonical row. Returns (rows, stats) in the exact
    FORM B column layout.
    """
    groups = {}
    for s in students:
        for c in s.courses:
            key = code_key(c.code)
            g = groups.setdefault(key, {
                "code": Counter(), "description": Counter(),
                "semester": Counter(), "credit": Counter(), "students": set(),
            })
            g["code"][str(c.code or "").strip()] += 1
            g["description"][str(c.title or "").strip()] += 1
            g["semester"][c.semester] += 1
            g["credit"][c.credit_value] += 1
            g["students"].add(s.matricule)

    rows = []
    for key, g in groups.items():
        rows.append({
            "code": g["code"].most_common(1)[0][0],
            "description": g["description"].most_common(1)[0][0],
            "credit": g["credit"].most_common(1)[0][0],
            "semester": g["semester"].most_common(1)[0][0],
            "department": (department or "").upper(),
            "levels": int(level) if str(level or "").isdigit() else 0,
            "school_id": school_id,
            "year": year or "",
            "section": section,
            "_students": sorted(g["students"]),
        })
    rows.sort(key=lambda r: (r["semester"], r["code"]))

    stats = {
        "catalog_count": len(rows),
        "student_count": len(students),
        "course_stats": [
            {**r, "taken_by": r["_students"], "count": len(r["_students"])}
            for r in rows
        ],
        "students": [
            {"matricule": s.matricule, "name": s.name, "taken": len(s.courses),
             "matched": len(s.courses), "missing": [], "unexpected": [],
             "issues": []}
            for s in students
        ],
        "unexpected": [],
    }
    return rows, stats


# ═══════════════════════════════════════════════════════════════════════════
#  EXCEL OUTPUT
# ═══════════════════════════════════════════════════════════════════════════
def _hcell(ws, r, c, v, fill=DARK_BLUE, fc="FFFFFF", sz=10):
    x = ws.cell(r, c, v)
    x.font = Font(name="Arial", bold=True, color=fc, size=sz)
    x.fill = PatternFill("solid", fgColor=fill)
    x.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    x.border = bdr
    return x


def _dcell(ws, r, c, v, fill=PALE_GREY, bold=False, ha="left"):
    x = ws.cell(r, c, v)
    x.font = Font(name="Arial", size=9, bold=bold)
    x.fill = PatternFill("solid", fgColor=fill)
    x.alignment = Alignment(horizontal=ha, vertical="center")
    x.border = bdr
    return x


def write_form_b_sheet(ws, catalog):
    """Write the canonical 9-column FORM B sheet."""
    for col, h in enumerate(FORM_B_HEADERS, 1):
        _hcell(ws, 1, col, h)
    ws.row_dimensions[1].height = 22
    ws.freeze_panes = "A2"

    for i, r in enumerate(catalog, 2):
        fill = SEM_FILLS.get(r["semester"], PALE_GREY)
        values = [
            r["code"], r["description"], r["credit"], r["semester"],
            r["department"], r["levels"], r["school_id"], r["year"], r["section"],
        ]
        for col, v in enumerate(values, 1):
            _dcell(ws, i, col, v, fill, ha="left" if col in (1, 2, 5) else "center")
        ws.row_dimensions[i].height = 14

    for col, w in enumerate([13, 42, 8, 10, 26, 8, 10, 12, 9], 1):
        ws.column_dimensions[get_column_letter(col)].width = w


def write_cross_check_sheet(ws, stats):
    """Snapshot of the cross-check: course coverage + per-student issues."""
    row = 1
    ws.cell(row, 1, "COURSE COVERAGE").font = Font(name="Arial", bold=True, size=11, color=DARK_BLUE)
    row += 1
    for col, h in enumerate(["Course Code", "Descriptions", "Semester", "Credit",
                             "Students", "Coverage %"], 1):
        _hcell(ws, row, col, h)
    row += 1
    total = max(stats["student_count"], 1)
    for cs in stats["course_stats"]:
        pct = round(cs["count"] / total * 100)
        for col, v in enumerate([cs["code"], cs["description"], cs["semester"],
                                 cs["credit"], cs["count"], pct], 1):
            _dcell(ws, row, col, v, ha="left" if col == 2 else "center")
        row += 1

    row += 1
    ws.cell(row, 1, "STUDENT ISSUES").font = Font(name="Arial", bold=True, size=11, color=DARK_BLUE)
    row += 1
    for col, h in enumerate(["Matricule", "Name", "Taken", "Matched",
                             "Missing", "Unexpected", "Credit/Sem issues"], 1):
        _hcell(ws, row, col, h)
    row += 1
    for ps in stats["students"]:
        fill = RED_FILL if (ps["missing"] or ps["unexpected"] or ps["issues"]) else PALE_GREEN
        for col, v in enumerate([ps["matricule"], ps["name"], ps["taken"], ps["matched"],
                                 len(ps["missing"]), len(ps["unexpected"]),
                                 len(ps["issues"])], 1):
            _dcell(ws, row, col, v, fill, ha="left" if col == 2 else "center")
        row += 1

    for col, w in enumerate([18, 32, 8, 9, 9, 11, 15], 1):
        ws.column_dimensions[get_column_letter(col)].width = w


def build_form_b_excel(catalog, stats, output_path):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Form B"
    write_form_b_sheet(ws, catalog)

    ws2 = wb.create_sheet("Cross-Check")
    write_cross_check_sheet(ws2, stats)

    wb.save(str(output_path))
    return str(output_path)


def add_form_b_sheet(wb, catalog):
    if "Form B" in wb.sheetnames:
        del wb["Form B"]
    ws = wb.create_sheet("Form B")
    write_form_b_sheet(ws, catalog)
    return ws
