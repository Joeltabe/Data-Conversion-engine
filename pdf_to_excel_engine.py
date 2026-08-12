#!/usr/bin/env python3
"""
=============================================================================
  ACADEMIC TRANSCRIPT PDF → EXCEL CONVERSION ENGINE  v2.0
  Generalised — any class, any number of courses, any level, any specialty
=============================================================================
USAGE (standalone):
    python pdf_to_excel_engine.py results.pdf
    python pdf_to_excel_engine.py results.pdf -o out.xlsx --school-id 10
    python pdf_to_excel_engine.py a.pdf b.pdf c.pdf --batch -o merged.xlsx
=============================================================================
"""

import sys, os, re, argparse, traceback
from pathlib import Path
from datetime import datetime

try:
    import pdfplumber
except ImportError:
    print("ERROR: pdfplumber not installed. Run: pip install pdfplumber"); sys.exit(1)

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    print("ERROR: openpyxl not installed. Run: pip install openpyxl"); sys.exit(1)

DEFAULT_SCHOOL_ID = 10

SEMESTER_MARKERS = {
    1: ["first semester","semester 1","semester i","1st semester"],
    2: ["second semester","semester 2","semester ii","2nd semester"],
    3: ["third semester","semester 3","semester iii","3rd semester"],
    4: ["fourth semester","semester 4","semester iv","4th semester"],
}

SKIP_ROW_MARKERS = [
    "credit value","credit earned","total","grade system","grade mark",
    "overall credit","gpa","cummulative","cumulative","dr.","registrar",
    "tel:","email:","website:","course code","course title",
]

COL_CODE  = ["course code","code"]
COL_TITLE = ["course title","title"]
COL_CA    = ["ca / 30","ca/30","ca","test","ca30"]
COL_EXAM  = ["exam / 70","exam/70","exam","exam70"]
COL_TOTAL = ["total / 100","total/100","total"]

# ── Grade bands (as printed on the transcripts) ─────────────────────────────
GRADE_BANDS = {"A":(80,100),"B+":(70,79),"B":(60,69),"C+":(55,59),
               "C":(50,54),"D+":(45,49),"D":(40,44),"F":(0,39)}

def _grade_band(total):
    for g,(lo,hi) in GRADE_BANDS.items():
        if lo<=total<=hi: return g
    return None

# ── Colours ──────────────────────────────────────────────────────────────────
DARK_BLUE  = "1F4E79"; MID_BLUE = "2E75B6"; PALE_BLUE = "EBF3FB"
PALE_YELL  = "FFF2CC"; PALE_GREY = "F8F9FA"; PALE_GREEN = "E2EFDA"
RED_FILL   = "FFE0E0"; TITLE_BG = "D6E4F0"; INFO_BG = "EEF5FB"
GREEN_TXT  = "375623"; RED_TXT = "C00000"
thin = Side(style="thin", color="BFBFBF")
med  = Side(style="medium", color="2E75B6")
bdr  = Border(left=thin, right=thin, top=thin, bottom=thin)


# ═══════════════════════════════════════════════════════════════════════════
#  LOGGER
# ═══════════════════════════════════════════════════════════════════════════
class Logger:
    def __init__(self): self.warnings=[]; self.errors=[]
    def section(self, t): print(f"\n{'-'*56}\n  {t}\n{'-'*56}")
    def step(self, n, t, m): print(f"  [{'▓'*n}{'░'*(t-n)}] {n}/{t}  {m}")
    def ok(self, m):   print(f"    ✓  {m}")
    def warn(self, m): print(f"    ⚠  {m}"); self.warnings.append(m)
    def err(self, m):  print(f"    ✗  {m}"); self.errors.append(m)
    def info(self, m): print(f"       {m}")
    def summary(self):
        print(f"\n{'='*56}\n  Warnings:{len(self.warnings)}  Errors:{len(self.errors)}")
        for w in self.warnings: print(f"    ⚠ {w}")
        for e in self.errors:   print(f"    ✗ {e}")

log = Logger()


# ═══════════════════════════════════════════════════════════════════════════
#  DATA STRUCTURES
# ═══════════════════════════════════════════════════════════════════════════
class CourseRecord:
    __slots__ = ("code","title","status","credit_value","credit_earned",
                 "ca","exam","total","grade_point","weighted","grade","semester",
                 "total_only")
    def __init__(self,code,title,status,credit_value,credit_earned,
                 ca,exam,total,grade_point,weighted,grade,semester,total_only=False):
        self.code=code; self.title=title; self.status=status
        self.credit_value=credit_value; self.credit_earned=credit_earned
        self.ca=ca; self.exam=exam; self.total=total
        self.grade_point=grade_point; self.weighted=weighted
        self.grade=grade; self.semester=semester
        self.total_only=total_only

    def validation_errors(self):
        e=[]
        if not self.code: e.append("missing subject code")
        if not (0<=self.ca<=30):   e.append(f"CA={self.ca} out of range [0–30]")
        exam_max = 100 if self.total_only else 70
        if not (0<=self.exam<=exam_max): e.append(f"Exam={self.exam} out of range [0–{exam_max}]")
        computed = round(self.ca+self.exam,2)
        if abs(computed-self.total)>0.5:
            e.append(f"Total mismatch: {self.ca}+{self.exam}={computed} ≠ {self.total}")
        gb = _grade_band(self.total)
        if gb and self.grade and gb != self.grade.upper():
            e.append(f"Grade {self.grade} doesn't match total {self.total} (expected {gb})")
        if self.grade.upper()=="F" and self.credit_earned!=0:
            e.append(f"Failed course but credit earned={self.credit_earned}")
        if self.grade.upper()!="F" and self.credit_earned==0 and self.total>0:
            e.append(f"Passing course ({self.total}) with credit earned=0")
        return e


class StudentRecord:
    def __init__(self,matricule,name,faculty="",specialty="",
                 department="",level="",academic_year=""):
        self.matricule     = re.sub(r'\s+','',matricule.strip())
        self.name          = name.strip()
        self.faculty       = faculty
        self.specialty     = specialty
        self.department    = department
        self.level         = level
        self.academic_year = academic_year
        self.courses       = []
        self.source_page   = None
        self.parse_warnings= []

    def validation_errors(self):
        e=[]
        if not self.matricule: e.append("missing matricule")
        if not self.name:      e.append("missing name")
        if not self.courses:   e.append("no courses found")
        for c in self.courses:
            for ce in c.validation_errors():
                e.append(f"{c.code} Sem{c.semester}: {ce}")
        return e


# ═══════════════════════════════════════════════════════════════════════════
#  HELPERS
# ═══════════════════════════════════════════════════════════════════════════
def _f(val, default=0.0):
    try: return float(str(val).replace(",",".").strip())
    except: return default

def _norm_code(val):
    """Course codes arrive with stray spaces (e.g. 'ACC 17') and trailing
    punctuation (e.g. 'CET14.') — strip them."""
    code = re.sub(r"\s+", "", str(val or "")).upper()
    return code.rstrip(".").rstrip()

def _is_skip(row):
    if not row or not row[0]: return True
    f = str(row[0]).strip().lower()
    return any(m in f for m in SKIP_ROW_MARKERS)

def _col_map(header):
    m={}
    for i,c in enumerate(header):
        if c is None: continue
        h=re.sub(r"\s+", " ", str(c).strip().lower())
        for n in COL_CODE:
            if n in h: m.setdefault("code",i)
        for n in COL_TITLE:
            if n in h: m.setdefault("title",i)
        if "status"==h: m.setdefault("status",i)
        if "credit value" in h: m.setdefault("credit_value",i)
        if "credit earned" in h: m.setdefault("credit_earned",i)
        for n in COL_CA:
            if n in h: m.setdefault("ca",i)
        for n in COL_EXAM:
            if n in h: m.setdefault("exam",i)
        for n in COL_TOTAL:
            if n in h and "grade" not in h: m.setdefault("total",i)
        if "grade point" in h or h=="grade\npoint": m.setdefault("grade_point",i)
        if "weighted" in h: m.setdefault("weighted",i)
        if h=="grade": m.setdefault("grade",i)
    return m

def _is_course_hdr(row):
    if not row: return False
    norm=" ".join(re.sub(r"\s+"," ",str(c).lower()) for c in row if c)
    return any(n in norm for n in COL_CODE)

def _sem_from_text(text):
    t=text.lower()
    for n,markers in SEMESTER_MARKERS.items():
        for m in markers:
            if m in t: return n
    return None


# ═══════════════════════════════════════════════════════════════════════════
#  PDF PARSER
# ═══════════════════════════════════════════════════════════════════════════
def _student_header(text):
    info={}
    # Value stops at the next known field label (e.g. "LEVEL:", "DATE AND
    # PLACE OF BIRTH:"), at 2+ spaces, a newline, or end of text.
    _labels = (r"(?:NAME AND SURNAME|DATE AND PLACE OF\s*BIRTH|FACULTY|DEPARTMENT|"
               r"SPECIALTY|PROGRAMME|PROGRAM|MATRICULE|GENDER|LEVEL|DATE|DEPT)")
    stop = rf"(?=\s+{_labels}\b\s*:|\s{{2,}}|[\r\n]|$)"

    # Name — handles the standard "NAME AND SURNAME: X" form plus the
    # multi-line layouts seen on results sheets where the label is split
    # ("NAME AND : X" + "SURNAME Y") and where "DATE AND PLACE OF BIRTH:"
    # is broken across lines. The name always stops before the DOB / the
    # next header field, so stray date text is never captured.
    name_block = None
    m = re.search(
        r"name\s+and\s*(?:surname)?\s*[:]?\s*"
        r"(.*?)"
        r"(?=\s*(?:date\s+and\s+place|matricule|gender|faculty|department|"
        r"specialty|level|programme|program|results\s+for|course\s+code|$))",
        text, re.IGNORECASE | re.DOTALL)
    if m:
        name_block = m.group(1)
        name_block = re.sub(r"\b(?:surname|name\s+and)\b", " ", name_block,
                            flags=re.IGNORECASE)
        info["name"] = " ".join(name_block.split())

    pats={
        "matricule": r"matricule[:\s]+([A-Z0-9][A-Z0-9\-\s]{0,8}[A-Z0-9]+?)(?=\s{2,}|\s*(?:GENDER|FACULTY|DATE|DEPT|LEVEL|$))",
        "faculty":   r"faculty[:\s]+([^\n\r]+?)" + stop,
        "specialty": r"specialty[:\s]+([^\n\r]+?)" + stop,
        "department":r"department[:\s]+([^\n\r]+?)" + stop,
        "level":     r"level[:\s]+(\d+)",
        "year":      r"(?:academic transcript for|results\s+for)\s+([\d/]+)",
    }
    for k,p in pats.items():
        m=re.search(p,text,re.IGNORECASE)
        if m: info[k]=m.group(1).strip()
    return info

def parse_pdf(pdf_path, school_id=DEFAULT_SCHOOL_ID):
    students=[]
    try:
        pdf=pdfplumber.open(str(pdf_path))
    except Exception as e:
        log.err(f"Cannot open PDF: {e}"); return students

    log.info(f"Pages: {len(pdf.pages)}")
    for page_num, page in enumerate(pdf.pages,1):
        try:
            raw=page.extract_text() or ""
        except Exception as e:
            log.warn(f"Page {page_num}: text extraction failed — {e}"); raw=""

        if "matricule" not in raw.lower():
            log.info(f"  Page {page_num}: skip (no student)"); continue

        hdr=_student_header(raw)
        mat=hdr.get("matricule",""); name=hdr.get("name","")
        if not mat:
            log.warn(f"  Page {page_num}: no matricule found ⚠️ — will convert with placeholder")

        s=StudentRecord(mat,name,hdr.get("faculty",""),hdr.get("specialty",""),
                        hdr.get("department",""),hdr.get("level",""),hdr.get("year",""))
        s.source_page=page_num

        try: tables=page.extract_tables()
        except: tables=[]

        if tables:
            _parse_tables(tables, raw, s)
        else:
            log.warn(f"  Page {page_num}: no tables, using text fallback")
            _parse_text(raw, s)

        if s.courses:
            students.append(s)
            log.ok(f"  Page {page_num}: {mat} — {name} — {len(s.courses)} courses")
        else:
            log.warn(f"  Page {page_num}: {mat} — 0 courses extracted")
            students.append(s)

    pdf.close()
    return students

def _parse_tables(tables, raw_text, student):
    course_tables=[t for t in tables if t and len(t[0])>=6 and _is_course_hdr(t[0])]
    text_lower=raw_text.lower()
    sem_pos=[]
    for n,markers in SEMESTER_MARKERS.items():
        for m in markers:
            idx=text_lower.find(m)
            if idx!=-1: sem_pos.append((idx,n))
    sem_pos.sort()
    if not sem_pos: sem_pos=[(0,1)]

    n_sems=len(set(s for _,s in sem_pos))
    tps=max(1,len(course_tables)//n_sems) if n_sems else 1

    for ti,table in enumerate(course_tables):
        sem_nums=sorted(set(s for _,s in sem_pos))
        sem=sem_nums[min(ti//tps, len(sem_nums)-1)]
        cm=_col_map(table[0])
        if "code" not in cm: continue

        for row in table[1:]:
            if _is_skip(row): continue
            if len(row)<=max(cm.values()): continue
            code=_norm_code(row[cm["code"]]) if cm.get("code") is not None else ""
            if not code or not re.match(r"^[A-Z]{2,6}\d{0,4}$",code): continue

            total_only = cm.get("total") is not None and cm.get("ca") is None and cm.get("exam") is None
            if total_only:
                total = _f(row[cm["total"]])
                ca    = round(total*0.3, 2)
                exam  = round(total - ca, 2)
            else:
                ca    = _f(row[cm["ca"]])    if cm.get("ca")    is not None else 0.0
                exam  = _f(row[cm["exam"]])  if cm.get("exam")  is not None else 0.0
                total = _f(row[cm["total"]]) if cm.get("total") is not None else round(ca+exam,2)
                if total==0 and (ca+exam)>0: total=round(ca+exam,2)

            student.courses.append(CourseRecord(
                code=code,
                title=str(row[cm["title"]]).strip() if cm.get("title") is not None else "",
                status=str(row[cm.get("status",0)]).strip() if cm.get("status") is not None else "",
                credit_value  =_f(row[cm["credit_value"]])   if cm.get("credit_value")  is not None else 0.0,
                credit_earned =_f(row[cm["credit_earned"]])  if cm.get("credit_earned") is not None else 0.0,
                ca=ca, exam=exam, total=total,
                grade_point=_f(row[cm["grade_point"]]) if cm.get("grade_point") is not None else 0.0,
                weighted   =_f(row[cm["weighted"]])    if cm.get("weighted")    is not None else 0.0,
                grade=str(row[cm["grade"]]).strip()    if cm.get("grade")       is not None else "",
                semester=sem,
                total_only=total_only
            ))

def _parse_text(raw, student):
    lines=raw.split("\n"); cur=1
    crx=re.compile(
        r"^([A-Z]{2,6}\d{0,4}\.?\b)\s+(.+?)\s+([CE])\s+"
        r"(\d+\.\d+)\s+(\d+\.\d+)\s+(\d+\.\d+)\s+(\d+\.\d+)\s+"
        r"(\d+\.\d+)\s+(\d+\.\d+)\s+(\d+\.\d+)\s+([A-F][+]?)$"
    )
    crx_only=re.compile(
        r"^([A-Z]{2,6}\d{0,4}\.?\b)\s+(.+?)\s+([CE])\s+"
        r"(\d+\.\d+)\s+(\d+\.\d+)\s+(\d+\.\d+)\s+(\d+\.\d+)\s+(\d+\.\d+)\s+([A-F][+]?)$"
    )
    for line in lines:
        line=line.strip()
        s=_sem_from_text(line)
        if s: cur=s; continue
        m=crx.match(line)
        if m:
            student.courses.append(CourseRecord(
                code=_norm_code(m.group(1)),title=m.group(2).strip(),status=m.group(3),
                credit_value=float(m.group(4)),credit_earned=float(m.group(5)),
                ca=float(m.group(6)),exam=float(m.group(7)),total=float(m.group(8)),
                grade_point=float(m.group(9)),weighted=float(m.group(10)),
                grade=m.group(11),semester=cur
            ))
            continue
        m=crx_only.match(line)
        if m:
            total=float(m.group(6))
            ca=round(total*0.3,2)
            student.courses.append(CourseRecord(
                code=_norm_code(m.group(1)),title=m.group(2).strip(),status=m.group(3),
                credit_value=float(m.group(4)),credit_earned=float(m.group(5)),
                ca=ca,exam=round(total-ca,2),total=total,
                grade_point=float(m.group(7)),weighted=float(m.group(8)),
                grade=m.group(9),semester=cur,total_only=True
            ))


# ═══════════════════════════════════════════════════════════════════════════
#  PDF SPLITTER  (combined Results PDF → one PDF per student)
# ═══════════════════════════════════════════════════════════════════════════
def _safe_filename(txt, max_len=60):
    txt = re.sub(r"[^\w\-. ]+", "", txt or "").strip().replace(" ", "_")
    return (txt or "STUDENT")[:max_len]

def split_results_pdf(pdf_path, out_dir, school_id=DEFAULT_SCHOOL_ID):
    """Split a combined results PDF (one student per page) into individual
    per-student PDFs. Returns list of (matricule, name, filename, source_page)."""
    from pypdf import PdfReader, PdfWriter
    out_dir = Path(out_dir); out_dir.mkdir(parents=True, exist_ok=True)

    written = []
    try:
        reader = PdfReader(str(pdf_path))
        with pdfplumber.open(str(pdf_path)) as pdf:
            for page_num, page in enumerate(pdf.pages, 1):
                try: raw = page.extract_text() or ""
                except Exception: raw = ""
                if "matricule" not in raw.lower():
                    log.info(f"  Page {page_num}: skip (no student)"); continue

                hdr = _student_header(raw)
                mat = hdr.get("matricule", "").replace(" ", "")
                name = hdr.get("name", "")
                if not mat:
                    log.warn(f"  Page {page_num}: no matricule — skip split"); continue

                base = f"{mat}_{_safe_filename(name)}" if mat else _safe_filename(name)
                fname = f"{base}.pdf"; out = out_dir / fname

                writer = PdfWriter()
                writer.add_page(reader.pages[page_num - 1])
                with open(out, "wb") as f:
                    writer.write(f)
                written.append((mat, name, fname, page_num))
                log.ok(f"  Page {page_num}: {mat} — {name} → {fname}")
    except Exception as e:
        log.err(f"Split failed: {e}")
    return written


# ═══════════════════════════════════════════════════════════════════════════
#  VALIDATION
# ═══════════════════════════════════════════════════════════════════════════
def validate_all(students):
    passed=0; failed=0; errors={}
    for s in students:
        errs=s.validation_errors()
        if errs:
            failed+=1; errors[s.matricule]=errs
            log.warn(f"{s.matricule} — {len(errs)} issue(s):")
            for e in errs[:4]: log.info(f"  → {e}")
        else:
            passed+=1; log.ok(f"{s.matricule} — {s.name} — PASS")
    return passed, failed, errors


# ═══════════════════════════════════════════════════════════════════════════
#  EXCEL BUILDER
# ═══════════════════════════════════════════════════════════════════════════
def _hcell(ws,r,c,v,fill,fc="FFFFFF",bold=True,sz=10,wrap=False,ha="center"):
    x=ws.cell(r,c,v)
    x.font=Font(name="Arial",bold=bold,color=fc,size=sz)
    x.fill=PatternFill("solid",fgColor=fill)
    x.alignment=Alignment(horizontal=ha,vertical="center",wrap_text=wrap)
    x.border=bdr; return x

def _dcell(ws,r,c,v,fill=PALE_BLUE,bold=False,color="000000",ha="center"):
    x=ws.cell(r,c,v)
    x.font=Font(name="Arial",size=9,bold=bold,color=color)
    x.fill=PatternFill("solid",fgColor=fill)
    x.alignment=Alignment(horizontal=ha,vertical="center")
    x.border=bdr; return x

def build_excel(students, output_path, school_id=DEFAULT_SCHOOL_ID, form_b_rows=None):
    wb=openpyxl.Workbook(); ws=wb.active; ws.title="Results"

    specialty  = students[0].specialty    if students else ""
    level      = students[0].level        if students else ""
    acad_year  = students[0].academic_year if students else ""
    dept       = students[0].department   if students else ""
    title_txt  = f"STUDENT RESULTS — {specialty.upper() or 'ACADEMIC'} LEVEL {level}  |  {acad_year}"

    # ── Flat, import-ready Results sheet ───────────────────────────────────
    hdrs=["Matricule","Subject Code","Schoolid","Test","Exam","Db1/Acc","Semester"]
    for col,h in enumerate(hdrs,1): _hcell(ws,1,col,h,DARK_BLUE,wrap=True)
    ws.row_dimensions[1].height=22; ws.freeze_panes="A2"

    sem_fills={1:PALE_BLUE,2:PALE_YELL,3:"E8F5E9",4:"F3E5F5"}
    row_idx=2

    for student in students:
        if not student.courses: continue
        for course in student.courses:
            fill=sem_fills.get(course.semester,PALE_GREY)
            if course.validation_errors(): fill=RED_FILL
            mf="FFF2CC" if not student.matricule else PALE_GREY

            _dcell(ws,row_idx,1,student.matricule,mf,ha="left")
            _dcell(ws,row_idx,2,course.code,   fill,ha="left")
            _dcell(ws,row_idx,3,school_id,     fill)
            _dcell(ws,row_idx,4,course.ca,     fill)
            _dcell(ws,row_idx,5,course.exam,   fill)
            _dcell(ws,row_idx,6,student.academic_year,fill)
            _dcell(ws,row_idx,7,course.semester,fill)
            ws.row_dimensions[row_idx].height=14; row_idx+=1

    for col,w in enumerate([18,14,10,8,8,12,10],1):
        ws.column_dimensions[get_column_letter(col)].width=w

    # ── Summary sheet ──────────────────────────────────────────────────────
    ws2=wb.create_sheet("Summary")
    ws2.merge_cells("A1:H1")
    c=ws2["A1"]; c.value=f"STUDENT SUMMARY — {title_txt}"
    c.font=Font(name="Arial",bold=True,size=11,color=DARK_BLUE)
    c.alignment=Alignment(horizontal="center",vertical="center")
    c.fill=PatternFill("solid",fgColor=TITLE_BG); ws2.row_dimensions[1].height=20

    s2h=["Matricule","Name","Level","Specialty","Sem 1","Sem 2","Total Courses","Validation"]
    for col,h in enumerate(s2h,1): _hcell(ws2,2,col,h,DARK_BLUE,wrap=True)
    ws2.row_dimensions[2].height=22; ws2.freeze_panes="A3"

    for r,s in enumerate(students,3):
        errs=s.validation_errors()
        sf=PALE_YELL if not s.matricule else (PALE_GREEN if not errs else RED_FILL)
        sc="996600" if not s.matricule else (GREEN_TXT  if not errs else RED_TXT)
        sc_counts={};
        for c in s.courses: sc_counts[c.semester]=sc_counts.get(c.semester,0)+1
        row_data=[s.matricule,s.name,s.level,s.specialty,
                  sc_counts.get(1,0),sc_counts.get(2,0),len(s.courses),
                  "✓ PASS" if not errs else f"✗ {len(errs)} issue(s)"]
        bf=PALE_BLUE if r%2==0 else PALE_GREY
        for col,val in enumerate(row_data,1):
            f=sf if col==8 else bf
            fc=sc if col==8 else "000000"
            _dcell(ws2,r,col,val,f,bold=(col==8 and not errs),color=fc,
                   ha="left" if col<=4 else "center")
        ws2.row_dimensions[r].height=15

    for col,w in enumerate([18,30,8,16,8,8,14,20],1):
        ws2.column_dimensions[get_column_letter(col)].width=w

    # ── Legend sheet ───────────────────────────────────────────────────────
    ws3=wb.create_sheet("Legend")
    ws3["A1"].value="COLOUR LEGEND & COLUMN GUIDE"
    ws3["A1"].font=Font(name="Arial",bold=True,size=12,color=DARK_BLUE)
    for row,fill,desc in [
        (3,PALE_BLUE,"First Semester rows"),
        (4,PALE_YELL,"Second Semester rows"),
        (5,"E8F5E9", "Third Semester rows"),
        (6,RED_FILL, "Data error — value out of valid range"),
        (7,"FFF2CC", "No matricule found — manual review needed"),
        (8,PALE_GREY,"Alternating student separator"),
    ]:
        ws3.cell(row,1).fill=PatternFill("solid",fgColor=fill)
        ws3.cell(row,1).value=f"  {desc}"
        ws3.cell(row,1).font=Font(name="Arial",size=10)

    guides=[
        (10,"Matricule",         "Student registration number (repeated per course row)"),
        (11,"Subject Code",      "Course code (e.g. ACC202, FRE101)"),
        (12,"Schoolid",          f"Institution identifier — value: {school_id}"),
        (13,"Test",              "Continuous Assessment mark, out of 30"),
        (14,"Exam",              "Examination mark, out of 70"),
        (15,"Db1/Acc",           "Academic year (e.g. 2023/2024)"),
        (16,"Semester",          "1=First, 2=Second, 3=Third, etc."),
    ]
    ws3.cell(9,1,"COLUMN GUIDE").font=Font(name="Arial",bold=True,size=11,color=DARK_BLUE)
    for row,cn,desc in guides:
        ws3.cell(row,1,cn).font=Font(name="Arial",bold=True,size=10)
        ws3.cell(row,2,desc).font=Font(name="Arial",size=10)
    ws3.column_dimensions["A"].width=22; ws3.column_dimensions["B"].width=55

    if form_b_rows:
        try:
            from form_b import add_form_b_sheet
            add_form_b_sheet(wb, form_b_rows)
        except Exception as e:
            log.warn(f"Form B sheet skipped: {e}")

    wb.save(str(output_path))
    log.ok(f"Excel saved → {output_path}")


# ═══════════════════════════════════════════════════════════════════════════
#  CLI ENTRY POINT
# ═══════════════════════════════════════════════════════════════════════════
def main():
    ap=argparse.ArgumentParser(description="Convert transcript PDF(s) to Excel")
    ap.add_argument("inputs",nargs="+",help="PDF file(s)")
    ap.add_argument("--output","-o",default=None)
    ap.add_argument("--school-id","-s",type=int,default=DEFAULT_SCHOOL_ID)
    ap.add_argument("--split",action="store_true",
                    help="Also split each combined results PDF into per-student PDFs")
    ap.add_argument("--split-dir",default=None,
                    help="Output folder for split PDFs (default: <pdf>_split/)")
    ap.add_argument("--no-confirm",action="store_true")
    args=ap.parse_args()

    out=args.output or (Path(args.inputs[0]).stem+"_converted.xlsx" if len(args.inputs)==1
                        else f"merged_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx")
    print(f"\n{'='*56}\n  TRANSCRIPT ENGINE v2.0\n  Output: {out}\n{'='*56}")

    log.section("STEP 1/4 — Parsing")
    all_students=[]
    for p in args.inputs:
        if not os.path.isfile(p): log.err(f"Not found: {p}"); continue
        all_students.extend(parse_pdf(p, args.school_id))
    if not all_students: log.err("No students extracted."); sys.exit(1)

    if args.split:
        log.section("STEP 1b — Splitting combined PDFs")
        for p in args.inputs:
            if not os.path.isfile(p): continue
            split_dir = Path(args.split_dir) if args.split_dir else Path(p).parent / (Path(p).stem + "_split")
            log.info(f"Splitting: {Path(p).name} → {split_dir}")
            split_results_pdf(p, split_dir, args.school_id)

    log.section("STEP 2/4 — Validation")
    passed,failed,errors=validate_all(all_students)
    log.info(f"Passed:{passed}  Failed:{failed}")

    log.section("STEP 3/4 — Confirmation")
    if not args.no_confirm:
        print(f"\n  {len(all_students)} students, {sum(len(s.courses) for s in all_students)} rows, {failed} with errors")
        print("  Proceed? [Y/n] ", end="", flush=True)
        try:
            if input().strip().lower() not in ("","y","yes"): print("Aborted."); sys.exit(0)
        except (EOFError, KeyboardInterrupt): pass

    log.section("STEP 4/4 — Excel")
    try:
        build_excel(all_students, out, args.school_id)
    except Exception as e:
        log.err(f"Excel build failed: {e}"); traceback.print_exc(); sys.exit(1)

    log.summary()
    print(f"\n  ✅  Done → {out}\n")

if __name__=="__main__":
    main()