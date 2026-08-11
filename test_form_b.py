#!/usr/bin/env python3
"""Tests for form_b.py — catalogue loading, detection, cross-check, Excel.
Run: python test_form_b.py"""
import os
import sys
import tempfile
import unittest

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

from pdf_to_excel_engine import StudentRecord, CourseRecord  # noqa: E402
from form_b import (  # noqa: E402
    load_form_b, detect_department, majority_department,
    majority_specialty, infer_department,
    parse_catalog_filename, find_catalogs, code_key, normalize_code,
    cross_check, build_form_b_excel, derive_form_b_from_students,
)

BASE = os.path.dirname(os.path.abspath(__file__))
CATALOG = os.path.join(BASE, "form_b", "FORM B LEVEL 300 NURSING.xlsx")


def _stu(mat, courses):
    s = StudentRecord(mat, "TEST " + mat)
    for code, title, sem, credit in courses:
        s.courses.append(CourseRecord(
            code=code, title=title, status="", credit_value=credit,
            credit_earned=credit, ca=18, exam=60, total=78,
            grade_point=4, weighted=4 * credit, grade="A", semester=sem))
    return s


class TestLoadFormB(unittest.TestCase):
    def test_catalog_exists(self):
        self.assertTrue(os.path.isfile(CATALOG),
                        "FORM B LEVEL 300 NURSING.xlsx not in form_b/")

    def test_load_form_b(self):
        rows = load_form_b(CATALOG)
        self.assertEqual(len(rows), 28)
        r0 = rows[0]
        self.assertEqual(r0["code"], "NUR 201")
        self.assertEqual(r0["description"], "MEDICAL PATHOLOGY I")
        self.assertEqual(r0["credit"], 3.0)
        self.assertEqual(r0["semester"], 1)
        self.assertEqual(r0["department"], "NURSING")
        self.assertEqual(r0["levels"], 300)
        self.assertEqual(r0["school_id"], 10)
        self.assertEqual(r0["year"], "2025/2026")
        self.assertEqual(r0["section"], 1)
        self.assertEqual(sorted({r["semester"] for r in rows}), [1, 2])


class TestDetection(unittest.TestCase):
    def test_detect_department(self):
        self.assertEqual(detect_department("LMU-24NUR001"), "NURSING")
        self.assertEqual(detect_department("LMU24NUR018"), "NURSING")
        self.assertEqual(detect_department("LMU-ACC001"), "ACCOUNTING")
        self.assertEqual(detect_department("LMUI250741"), "")
        self.assertEqual(detect_department(""), "")

    def test_majority_department(self):
        mats = ["LMU-24NUR001", "LMU24NUR018", "LMU-24NUR022", "LMUI250741"]
        self.assertEqual(majority_department(mats), "NURSING")

    def test_majority_specialty(self):
        sts = [_stu("LMU-ACC001", []), _stu("LMU-ACC002", []), _stu("LMU-ACC003", [])]
        for s in sts:
            s.specialty = "ACCOUNTING"
        sts[1].specialty = "COMPUTER GRAPHICS AND WEB DESIGN"
        self.assertEqual(majority_specialty(sts), "ACCOUNTING")
        for s in sts:
            s.specialty = ""
        self.assertEqual(majority_specialty(sts), "")

    def test_infer_department_prefers_specialty(self):
        sts = [_stu("LMU-ACC001", []), _stu("LMU-ACC002", [])]
        for s in sts:
            s.specialty = "COMPUTER GRAPHICS AND WEB DESIGN"
        self.assertEqual(infer_department(sts), "COMPUTER GRAPHICS AND WEB DESIGN")

    def test_infer_department_falls_back_to_matricule(self):
        sts = [_stu("LMU-24NUR001", []), _stu("LMU24NUR018", [])]
        self.assertEqual(infer_department(sts), "NURSING")

    def test_parse_catalog_filename(self):
        dept, lvl = parse_catalog_filename("FORM B LEVEL 300 NURSING.xlsx")
        self.assertEqual(dept, "NURSING")
        self.assertEqual(lvl, 300)

    def test_find_catalogs_ranks_matching_first(self):
        catalogs = find_catalogs(os.path.join(BASE, "form_b"), "NURSING", 300)
        self.assertTrue(catalogs)
        self.assertEqual(catalogs[0]["department"], "NURSING")
        self.assertEqual(catalogs[0]["level"], 300)


class TestCodeKeys(unittest.TestCase):
    def test_normalize(self):
        self.assertEqual(normalize_code(" mbs  201 "), "MBS201")
        self.assertEqual(normalize_code("MBS212"), "MBS212")

    def test_fuzzy_key(self):
        self.assertEqual(code_key("MBS 201"), code_key("MBS201"))
        self.assertEqual(code_key("NUR201"), code_key("NUR2O1"))


class TestCrossCheck(unittest.TestCase):
    def test_full(self):
        catalog = load_form_b(CATALOG)
        students = [
            _stu("LMU-24NUR001", [
                ("NUR 201", "MEDICAL PATHOLOGY I", 1, 3),
                ("MBS 201", "CLINICAL PHARMACOLOGY", 1, 3),
                ("NUR 202", "MEDICAL PATHOLOGY II", 2, 3),
                ("XYZ 999", "SOME EXTRA COURSE", 1, 2),
            ]),
            _stu("LMU-24NUR002", [
                ("NUR 201", "MEDICAL PATHOLOGY I", 1, 3),
                ("NUR 202", "MEDICAL PATHOLOGY II", 2, 3),
                ("MBS212", "CLINICAL INTERNSHIP AND REPORT II", 2, 6),
            ]),
        ]
        stats = cross_check(students, catalog)

        self.assertEqual(stats["catalog_count"], 28)
        self.assertEqual(stats["student_count"], 2)

        cov = {cs["code"]: cs["count"] for cs in stats["course_stats"]}
        self.assertEqual(cov["NUR 201"], 2)
        self.assertEqual(cov["MBS 201"], 1)
        self.assertEqual(cov["NUR 203"], 0)

        p0 = stats["students"][0]
        self.assertIn("XYZ 999", [u["code"] for u in p0["unexpected"]])
        missing0 = [r["code"] for r in p0["missing"]]
        self.assertIn("NUR 203", missing0)
        self.assertEqual(len(missing0), 25)  # 28 - 3 matched

        p1 = stats["students"][1]
        # MBS212 IS in the catalogue so it must be matched, not unexpected
        self.assertNotIn("MBS212", [u["code"] for u in p1["unexpected"]])

    def test_credit_and_semester_mismatch(self):
        catalog = load_form_b(CATALOG)
        s = _stu("LMU-24NUR003", [
            ("NUR 201", "MEDICAL PATHOLOGY I", 2, 4),
        ])
        stats = cross_check([s], catalog)
        ps = stats["students"][0]
        types = [(i["type"], i["code"]) for i in ps["issues"]]
        self.assertIn(("semester_mismatch", "NUR 201"), types)
        self.assertIn(("credit_mismatch", "NUR 201"), types)


class TestDeriveFormB(unittest.TestCase):
    def test_dedupes_and_sorts(self):
        students = [
            _stu("LMU-24NUR001", [
                ("MBS 201", "CLINICAL PHARMACOLOGY", 1, 3),
                ("NUR 201", "MEDICAL PATHOLOGY I", 1, 3),
                ("NUR 202", "MEDICAL PATHOLOGY II", 2, 3),
            ]),
            _stu("LMU-24NUR002", [
                ("NUR201", "MEDICAL PATHOLOGY I", 1, 3),  # same course, OCR spacing
                ("MBS201", "CLINICAL PHARMACOLOGY", 1, 3),
            ]),
        ]
        rows, stats = derive_form_b_from_students(
            students, department="NURSING", level="200", year="2023/2024",
            school_id=10, section=1)
        codes = [r["code"] for r in rows]
        self.assertEqual(codes, ["MBS 201", "NUR 201", "NUR 202"])  # sem 1 sorted, then sem 2
        self.assertEqual(rows[0]["semester"], 1)
        self.assertEqual(rows[2]["semester"], 2)
        self.assertEqual(rows[0]["department"], "NURSING")
        self.assertEqual(rows[0]["levels"], 200)
        self.assertEqual(rows[0]["year"], "2023/2024")
        self.assertEqual(rows[0]["section"], 1)
        self.assertEqual(rows[0]["school_id"], 10)
        self.assertEqual(stats["catalog_count"], 3)
        self.assertEqual(stats["student_count"], 2)
        # NUR 201 taken by both students (fuzzy match NUR201 vs NUR 201)
        cov = {cs["code"]: cs["count"] for cs in stats["course_stats"]}
        self.assertEqual(cov["NUR 201"], 2)

    def test_canonical_credit_and_description(self):
        students = [
            _stu("LMU-24NUR001", [("NUR 201", "MEDICAL PATHOLOGY I", 1, 3)]),
            _stu("LMU-24NUR002", [("NUR 201", "MEDICAL PATHOLOGY I", 1, 3)]),
            _stu("LMU-24NUR003", [("NUR 201", "MEDICAL PATHOLOGY 1", 1, 2)]),  # typo, minority
        ]
        rows, _ = derive_form_b_from_students(students, department="NURSING")
        r = rows[0]
        self.assertEqual(r["description"], "MEDICAL PATHOLOGY I")  # majority
        self.assertEqual(r["credit"], 3.0)                        # majority


class TestExcel(unittest.TestCase):
    def test_build_form_b_excel(self):
        catalog = load_form_b(CATALOG)
        students = [_stu("LMU-24NUR001", [("NUR 201", "MEDICAL PATHOLOGY I", 1, 3)])]
        stats = cross_check(students, catalog)
        with tempfile.TemporaryDirectory() as td:
            out = os.path.join(td, "FORM_B_out.xlsx")
            build_form_b_excel(catalog, stats, out)
            self.assertTrue(os.path.isfile(out))

            import openpyxl
            wb = openpyxl.load_workbook(out, data_only=True)
            self.assertEqual(wb.sheetnames, ["Form B", "Cross-Check"])
            ws = wb["Form B"]
            hdr = [ws.cell(1, c).value for c in range(1, 10)]
            self.assertEqual(hdr, ["Course Code", "Descriptions", "credit", "Semester",
                                   "Department", "Levels", "School id", "year", "section"])
            row2 = [ws.cell(2, c).value for c in range(1, 10)]
            self.assertEqual(row2, ["NUR 201", "MEDICAL PATHOLOGY I", 3, 1,
                                    "NURSING", 300, 10, "2025/2026", 1])

    def test_add_form_b_sheet_to_results_workbook(self):
        from pdf_to_excel_engine import build_excel
        catalog = load_form_b(CATALOG)
        students = [_stu("LMU-24NUR001", [("NUR 201", "MEDICAL PATHOLOGY I", 1, 3)])]
        with tempfile.TemporaryDirectory() as td:
            out = os.path.join(td, "results.xlsx")
            build_excel(students, out, 10, form_b_rows=catalog)
            import openpyxl
            wb = openpyxl.load_workbook(out)
            self.assertIn("Form B", wb.sheetnames)
            ws = wb["Form B"]
            self.assertEqual(ws.cell(1, 1).value, "Course Code")
            self.assertEqual(ws.cell(2, 1).value, "NUR 201")


if __name__ == "__main__":
    unittest.main(verbosity=2)
